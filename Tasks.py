import datetime
import openpyxl
import tkinter as tk
from tkcalendar import *
from tkinter import ttk, messagebox
import subprocess





class Tasks:
    def __init__(self, root):
        self.root = root
        self.root.title("Tasks!")
        
        self.TKINTER_WIDGETS = {}
        self.TKINTER_DATA = {}
        self.APP_WIDTH = 1300
        self.APP_HEIGHT = 900
        self.DEFAULT_TIME = "11:00"

        self.sheet_path = r"db_Tasks.xlsx"

        self._width = 50
        self.pad_x = 20
        self.pad_y = 10
        self.stick = "ew"

        self.valuesList = []        

        self.style = ttk.Style(self.root)
        
        self.root.tk.call("source",r"forest-dark.tcl")
        self.style.theme_use("forest-dark")
        self.style.configure('TLabelframe.Label', foreground ='#00936a')
        self.style.configure('TLabelframe.Label', font=('courier', 12, 'bold'))
        self.style.configure("Treeview.Heading", foreground="#00936a",font=('courier', 12, 'bold')) # foreground="white"
        self.style.map("TButton",foreground=[("disabled", "#807777")])
        self.style.map("TEntry",foreground=[("disabled", "white")])

        self.image = tk.PhotoImage(file=r"cal.png")

        ## Createing a frame
        self.frame = ttk.Frame(self.root)
        self.frame.pack(expand=0)

        self.frame_compo = ttk.LabelFrame(self.frame, text="Create a Task")
        self.frame_compo.grid(row=0,column=0,padx=15,pady=25)

        ## Create Entries
        self.title = self.create_entry_element(self.frame_compo,"label","Task Name",0,0)     ## Title Entry
        self.message = self.create_entry_element(self.frame_compo,"text","Description",1,0)    ## Message TextBox
        self.cmd = self.create_entry_element(self.frame_compo,"label","Command",3,0)   ## CMD Entry

        self.title_  = tk.StringVar()
        self.title['textvariable'] = self.title_
        self.title_.trace_add('write', self._state)

        ## DATE / TIME
        self.frame_datetime_ = ttk.LabelFrame(self.frame_compo,text="Date & Time")
        self.frame_datetime_.grid(row=5,column=0,pady=15)

        ## Entry Date
        current_date_time = datetime.datetime.now().strftime('%d/%m/%Y')

        self.entry_date = ttk.Entry(master=self.frame_datetime_,justify="center")
        self.entry_date.insert(0, current_date_time)
        self.entry_date.configure(state=tk.DISABLED,width=10)
        self.entry_date.grid(row=0, column=1,padx=30, pady=15)

        # Entry Time
        self.entry_time = ttk.Entry(master=self.frame_datetime_,justify="center")
        self.entry_time.insert(0, self.DEFAULT_TIME)
        self.entry_time.configure(state=tk.DISABLED,width=10)
        self.entry_time.grid(row=0, column=2,padx=30, pady=15)

        # Button Select Date Time
        self.btn_select_date_time = ttk.Button(master=self.frame_datetime_, image=self.image,  command=self.select_date_time)
        self.btn_select_date_time.grid(row=0, column=3,padx=30, pady=15)

        self.btn_save_data = ttk.Button(master=self.frame_compo,text="Save",command=self.save_data,state=tk.DISABLED)
        self.btn_save_data.grid(row=8,column=0,sticky=self.stick,padx=self.pad_x,pady=self.pad_y)

        ## Create Table      
        self.frame_table = ttk.Frame(self.frame)
        self.frame_table.grid(row=0,column=1,padx=25,pady=25)

        self.frame_table_compo = ttk.LabelFrame(self.frame_table, text="Tasks Table")
        self.frame_table_compo.grid(row=0,column=0)

        self.scroll_bar = ttk.Scrollbar(self.frame_table_compo,orient="vertical")

        self.cols = ("Title","Message","CMD","Date","Time","Status")
        self.tree_view = ttk.Treeview(self.frame_table_compo,show="headings",columns=self.cols,height=20,yscrollcommand=self.scroll_bar.set)
        self.tree_view.column("Title",width=100,anchor="center")
        self.tree_view.column("Message",width=170,anchor="w")
        self.tree_view.column("CMD",width=70,anchor="center")
        self.tree_view.column("Date",width=70,anchor="center")
        self.tree_view.column("Time",width=70,anchor="center")
        self.tree_view.column("Status",width=70,anchor="center")
        
        self.tree_view.grid(row=0,column=0,padx=20,pady=10)
        self.scroll_bar.config(command=self.tree_view.yview)

        self.frame_under_table = ttk.Frame(self.frame_table_compo)
        self.frame_under_table.grid(row=1,column=0,padx=10,pady=10)

        self.create_task_btn = ttk.Button(self.frame_under_table,text="Create Task",command=self.create_task)
        self.create_task_btn.grid(row=0,column=0,padx=10,pady=10)

        self.run_task_btn = ttk.Button(self.frame_under_table,text="Run Task",command=self.run_task)
        self.run_task_btn.grid(row=0,column=1,padx=10,pady=10)

        self.delete_row_btn = ttk.Button(self.frame_under_table,text="Delete Task",command=self.delete_a_row)
        self.delete_row_btn.grid(row=0,column=2,padx=10,pady=10)

        self.load_table()

    def run_task(self):
        '''Running a Task by name'''
        if self.check_item_selected():
            selected_item = self.tree_view.focus()
            values = self.tree_view.item(selected_item)["values"]
            command = f"SchTasks /run /tn \"{values[0]}\""

            self.run_schtask_cmd(command)         
        return

    def run_schtask_cmd(self,commando):
        try:
            subprocess.check_output(commando,shell=True,stderr=subprocess.STDOUT)
            print("Done!")

        except subprocess.CalledProcessError as e:
            error_message = e.output.decode().strip()
            messagebox.showerror("Error!!",error_message)

    def create_task(self):
        if self.check_item_selected():
            selected_item = self.tree_view.focus()
            values = self.tree_view.item(selected_item)["values"]
            
            text_ = 'SchTasks /Create /SC daily /TN '
            title_ = values[0]
            text_1 = ' /TR "cmd.exe /c '
            command = values[2]
            time_ = values[4]
            date_ = f" /sd {values[3]}"

            Task_Schedual_Command = text_ + '"' + title_ + '"' + text_1 + command + '"' + f" /ST {time_}" + date_
            self.run_schtask_cmd(Task_Schedual_Command)
        return

    def _state(self,*_):
        '''Change btn state by just filling in the task name'''
        if self.title_.get():
            self.btn_save_data['state'] = 'normal'
        else:
            self.btn_save_data['state'] = 'disabled'

    def check_item_selected(self):
        if self.tree_view.focus():
            return True
        return False

    def check_Task_name_match(self):
        entry_text = self.title.get().lower()

        for item in self.tree_view.get_children():
            item_value = self.tree_view.item(item)["values"][0].lower()
            if entry_text == item_value:
                return False
        return True

    def save_data(self):
        if self.check_Task_name_match():
            title_ = self.title.get()
            message_ = self.message.get("1.0",tk.END)
            
            cmd_ = self.cmd.get()
            date_ = self.entry_date.get()
            time_ = self.entry_time.get()
            ##  ##  ##  ##  ##  ##  ##  ##
            path = self.sheet_path
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            add_row_values = [title_,message_,cmd_,date_,time_]
            sheet.append(add_row_values)
            workbook.save(path)
            
            #insert in tabel
            self.tree_view.insert('',tk.END,values=add_row_values)
        
            #empty felds
            self.title.delete(0,tk.END)
            self.message.delete("1.0", tk.END)
            self.cmd.delete(0,tk.END)
        else:
            messagebox.showerror("Cant Add Task","Task name should be unique!")
        
    def load_table(self):
        
        path = self.sheet_path
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.tree_view.heading(col_name,text=col_name)
        for value_tuple in list_values[1:]:
            self.tree_view.insert("",tk.END,values=value_tuple)

    def delete_a_row(self):

        '''Delete selected row'''
        if self.check_item_selected():
            selected_item = self.tree_view.focus()
            values = self.tree_view.item(selected_item)["values"]

            confirm = messagebox.askyesno("Confirmation", f"Are you sure you want to Delete the task '{values[0]}'?")

            if confirm:
        
                workbook = openpyxl.load_workbook(self.sheet_path)
                worksheeet = workbook.active
                search_value = values[0]
                self.tree_view.delete(selected_item)

                for row_index, row in enumerate(worksheeet.iter_rows(values_only=True), start=1):
                    if search_value in row:
                    
                        worksheeet.delete_rows(row_index)
                        workbook.save(self.sheet_path)
                print(f"Task '{values[0]}' deleted successfully")
                try:
                    
                    command = f'schtasks /delete /tn \"{values[0]}\" /F'
                    subprocess.run(command, shell=True, check=True)               
                
                except subprocess.CalledProcessError as e:
                    # Capture and show the error message in a messagebox
                    error_message = e.stderr
                    print(error_message)
        return

    def create_entry_element(self,parent_frame,type,text_value,row_num,col_num):
        frame_ = ttk.LabelFrame(parent_frame,text=text_value)
        frame_.grid(row=row_num,column=col_num,padx=self.pad_x,pady=self.pad_y)

        if type == "label":
            entry_ = tk.Entry(frame_,borderwidth=0,width=self._width)
            
        if type == "text":
            entry_ = tk.Text(frame_,borderwidth=0,width=self._width,height=10)
        entry_.grid(row=row_num,column=col_num,sticky=self.stick,padx=self.pad_x,pady=self.pad_y)
        return entry_  

    def select_date_time(self):
            
        # Disable Button Select EID
        self.btn_select_date_time.configure(state=tk.DISABLED)
        
        # Create Top Level Window
        self.top_level_date_time = tk.Toplevel(takefocus=True)
        self.top_level_date_time.title("Select Date / Time")
                
        # Size
        top_level_date_time_width = 400
        top_level_date_time_height = 400

        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        app_center_coordinate_x = (screen_width / 3) - (top_level_date_time_width / 2.5)
        app_center_coordinate_y = (screen_height / 3) - (top_level_date_time_height / 2.5)
        self.top_level_date_time.geometry(f"{top_level_date_time_width}x{top_level_date_time_height}+{int(app_center_coordinate_x)}+{int(app_center_coordinate_y)}")

        self.top_level_date_time.resizable(width=False, height=False)

        # Close 'X' Button
        self.top_level_date_time.protocol("WM_DELETE_WINDOW", self.post_top_level_select_date_time)

        self.top_level_date_time.grid_rowconfigure(1, weight=1)
        self.top_level_date_time.grid_columnconfigure(0, weight=1)

        # Frame Calendar
        frame_calendar = ttk.Frame(master=self.top_level_date_time)
        frame_calendar.grid(row=0, column=0, padx=15, pady=15, columnspan=3)

        # Label Date
        self.label_date =  ttk.Label(master=frame_calendar, text="- Select Date -")
        self.label_date.grid(row=0, column=0, padx=10, pady=5, columnspan=3, sticky='n')

        # Calendar
        self.CAL = Calendar(frame_calendar, selectmode='day', date_pattern='dd/mm/y', mindate=datetime.datetime.today())
        self.CAL.grid(row=0, column=0, padx=20, pady=30, columnspan=3, sticky='s')

        # Frame Time
        frame_time = ttk.Frame(master=self.top_level_date_time)
        frame_time.grid(row=1, column=0, padx=15, pady=5, columnspan=3)

        # Time
        # Label Time
        self.label_time =  ttk.Label(master=frame_time, text="Time")
        self.label_time.grid(row=0, column=0, padx=10, pady=10)
        
        # Hour Time
        self.spinbox_hours = ttk.Spinbox(frame_time, width=10, justify=tk.CENTER,from_=00,to=23, format="%02.0f")
        self.spinbox_hours.grid(row=0, column=1, padx=5)
        
        # Set Default Value
        self.string_var_hours = tk.StringVar()
        self.string_var_hours.set(self.DEFAULT_TIME.split(':')[0])
        self.spinbox_hours.config(textvariable=self.string_var_hours)

        # SpinBox Minutes
        self.spinbox_minutes = ttk.Spinbox(frame_time, width=10, justify=tk.CENTER,from_=00,to=59,format="%02.0f")
        self.spinbox_minutes.grid(row=0, column=2, padx=5)
        
        # TextVariable for Minutes
        self.string_var_minutes = tk.StringVar()
        self.string_var_minutes.set(self.DEFAULT_TIME.split(':')[1])
        self.spinbox_minutes.config(textvariable=self.string_var_minutes)

        # Button Select Date Time OK
        self.button_select_date_time_ok = ttk.Button(master=self.top_level_date_time, text="OK", command=self.update_date_time)
        self.button_select_date_time_ok.grid(row=2, column=0, padx=60, pady=15, sticky='sw')

        # Button Select Date Time Cancel
        self.button_select_date_time_cancel = ttk.Button(master=self.top_level_date_time, text="Cancel", command=self.post_top_level_select_date_time)
        self.button_select_date_time_cancel.grid(row=2, column=0, padx=60, pady=15, sticky='se')


    # Post Top Level Select Date Time
    def post_top_level_select_date_time(self):
        
        # Destroy Top Level Date Time
        self.top_level_date_time.destroy()
        self.btn_select_date_time.configure(state=tk.NORMAL)

    # Update Date Time
    def update_date_time(self):
        
        new_date = self.CAL.get_date()
        new_time = f'{self.spinbox_hours.get()}:{self.spinbox_minutes.get()}'

        # Destroy Top Level Widget
        self.post_top_level_select_date_time()

        # Update Date Entry
        self.entry_date.configure(state=tk.NORMAL)
        self.entry_date.delete(0, tk.END)
        self.entry_date.insert(0, new_date)
        self.entry_date.configure(state=tk.DISABLED)

        # Update Time Entry
        self.entry_time.configure(state=tk.NORMAL)
        self.entry_time.delete(0, tk.END)
        self.entry_time.insert(0, new_time)
        self.entry_time.configure(state=tk.DISABLED)


root = tk.Tk()
task = Tasks(root)
root.mainloop()